import io
import csv
from urllib.parse import quote
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from errors import BusinessError, ErrorCode, ActionType
from sqlalchemy.orm import Session
from database import get_db
from account_dep import get_account_id
import crud
from openpyxl import Workbook

router = APIRouter()


def _stream_excel(wb: Workbook, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = quote(f"{filename}.xlsx", safe='')
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )


def _stream_csv(rows: list[list], headers: list[str], filename: str):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    buf.seek(0)
    encoded = quote(f"{filename}.csv", safe='')
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )


@router.get("/products")
def export_products(format: str = "excel", search: str = None, category: str = None, account_id: int = Depends(get_account_id), db: Session = Depends(get_db)):
    _, products = crud.list_products(db, account_id, limit=10000, search=search, category=category)
    headers = ["ID", "编码", "名称", "分类", "单位", "进价", "售价", "库存", "预警线", "描述"]
    rows = [[p.id, p.sku, p.name, p.category, p.unit, p.purchase_price, p.sale_price,
             p.inventory.quantity if p.inventory else 0, p.min_stock, p.description] for p in products]
    filename = "商品列表"
    if format == "csv":
        return _stream_csv(rows, headers, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "商品列表"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col in [5, 6]:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = '#,##0.00'
    return _stream_excel(wb, filename)


@router.get("/inventory")
def export_inventory(format: str = "excel", alert_only: bool = False, account_id: int = Depends(get_account_id), db: Session = Depends(get_db)):
    _, items = crud.list_inventory(db, account_id, limit=10000, alert_only=alert_only)
    headers = ["ID", "编码", "名称", "分类", "单位", "当前库存", "预警线", "状态", "进价", "售价", "库存价值"]
    rows = []
    for inv in items:
        p = inv.product
        qty = inv.quantity if inv.quantity is not None else 0
        status = "负库存" if qty < 0 else ("不足" if qty < (p.min_stock if p else 0) else "正常")
        rows.append([inv.id, p.sku if p else "", p.name if p else "", p.category if p else "",
                     p.unit if p else "", qty, p.min_stock if p else 0, status,
                     p.purchase_price if p else 0, p.sale_price if p else 0,
                     max(qty, 0) * (p.purchase_price if p else 0)])
    filename = "库存清单"
    if format == "csv":
        return _stream_csv(rows, headers, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "库存清单"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return _stream_excel(wb, filename)


@router.get("/purchases")
def export_purchases(format: str = "excel", start_date: str = None, end_date: str = None, account_id: int = Depends(get_account_id), db: Session = Depends(get_db)):
    orders = crud.get_purchase_report(db, account_id, start_date, end_date)
    headers = ["单号", "订单ID", "日期", "供应商", "订单总价", "状态", "商品", "数量", "单价", "小计", "备注"]
    rows = []
    total_qty = 0
    total_amount = 0
    for o in orders:
        for item in o.items:
            rows.append([
                o.order_no or "", o.id,
                o.purchase_date.strftime("%Y-%m-%d %H:%M") if o.purchase_date else "",
                o.supplier.name if o.supplier else "",
                o.total_price, o.status,
                item.product.name if item.product else "", item.quantity,
                item.unit_price, item.total_price, o.notes
            ])
            total_qty += item.quantity
            total_amount += item.total_price
    filename = "采购报表"
    if format == "csv":
        return _stream_csv(rows, headers, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "采购报表"
    ws.append(headers)
    ws.append(["", "", "", "合计", sum(o.total_price for o in orders), "", total_qty, "", total_amount, ""])
    for row in rows:
        ws.append(row)
    return _stream_excel(wb, filename)


@router.get("/sales")
def export_sales(format: str = "excel", start_date: str = None, end_date: str = None, account_id: int = Depends(get_account_id), db: Session = Depends(get_db)):
    orders = crud.get_sale_report(db, account_id, start_date, end_date)
    headers = ["单号", "订单ID", "日期", "客户", "订单总价", "状态", "商品", "数量", "单价", "小计", "备注"]
    rows = []
    total_qty = 0
    total_amount = 0
    for o in orders:
        for item in o.items:
            rows.append([
                o.order_no or "", o.id,
                o.sale_date.strftime("%Y-%m-%d %H:%M") if o.sale_date else "",
                o.customer.name if o.customer else "散客",
                o.total_price, o.status,
                item.product.name if item.product else "", item.quantity,
                item.unit_price, item.total_price, o.notes
            ])
            total_qty += item.quantity
            total_amount += item.total_price
    filename = "销售报表"
    if format == "csv":
        return _stream_csv(rows, headers, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "销售报表"
    ws.append(headers)
    ws.append(["", "", "", "合计", sum(o.total_price for o in orders), "", total_qty, "", total_amount, ""])
    for row in rows:
        ws.append(row)
    return _stream_excel(wb, filename)


@router.get("/products-batch")
def export_products_batch(
    product_ids: str,
    format: str = "excel",
    account_id: int = Depends(get_account_id),
    db: Session = Depends(get_db)
):
    """批量导出指定商品
    product_ids: 逗号分隔的商品ID列表，如 "1,2,3"
    """
    import models
    try:
        ids = [int(x.strip()) for x in product_ids.split(",") if x.strip()]
    except ValueError:
        raise BusinessError(code=ErrorCode.VALIDATION_ERROR, message="product_ids 必须是逗号分隔的整数列表")
    if not ids:
        raise BusinessError(code=ErrorCode.VALIDATION_ERROR, message="product_ids 不能为空")

    products = db.query(models.Product).filter(
        models.Product.account_id == account_id,
        models.Product.id.in_(ids)
    ).all()

    headers = ["ID", "编码", "名称", "分类", "单位", "进价", "售价", "库存", "预警线", "描述", "创建时间"]
    rows = []
    for p in products:
        rows.append([
            p.id, p.sku, p.name, p.category, p.unit,
            p.purchase_price, p.sale_price,
            p.inventory.quantity if p.inventory else 0,
            p.min_stock, p.description,
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""
        ])

    filename = f"库存商品导出_{len(products)}条"
    if format == "csv":
        return _stream_csv(rows, headers, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "商品导出"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col in [6, 7]:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = '#,##0.00'
    return _stream_excel(wb, filename)


@router.get("/profit")
def export_profit(format: str = "excel", start_date: str = None, end_date: str = None, account_id: int = Depends(get_account_id), db: Session = Depends(get_db)):
    data = crud.get_profit_report(db, account_id, start_date, end_date)
    purchase_orders = data.get("purchase_orders", [])
    sale_orders = data.get("sale_orders", [])
    wb = Workbook()
    ws = wb.active
    ws.title = "利润报表"
    ws.append(["采购汇总"])
    ws.append(["采购总金额", data.get("total_purchase_amount", 0), "采购笔数", data.get("purchase_count", 0)])
    ws.append(["销售总金额", data.get("total_sale_amount", 0), "销售笔数", data.get("sale_count", 0)])
    ws.append(["利润", data.get("total_profit", 0)])
    ws.append([])
    ws.append(["采购明细"])
    ws.append(["单号", "订单ID", "日期", "供应商", "商品", "数量", "单价", "小计", "状态"])
    for o in purchase_orders:
        for item in o.items:
            ws.append([o.order_no or "", o.id, o.purchase_date.strftime("%Y-%m-%d %H:%M") if o.purchase_date else "",
                       o.supplier.name if o.supplier else "", item.product.name if item.product else "",
                       item.quantity, item.unit_price, item.total_price, o.status])
    ws.append([])
    ws.append(["销售明细"])
    ws.append(["单号", "订单ID", "日期", "客户", "商品", "数量", "单价", "小计", "状态"])
    for o in sale_orders:
        for item in o.items:
            ws.append([o.order_no or "", o.id, o.sale_date.strftime("%Y-%m-%d %H:%M") if o.sale_date else "",
                       o.customer.name if o.customer else "散客", item.product.name if item.product else "",
                       item.quantity, item.unit_price, item.total_price, o.status])
    filename = "利润报表"
    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
        buf.seek(0)
        return StreamingResponse(buf, media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(f'{filename}.csv', safe='')}"})
    return _stream_excel(wb, filename)