"""集成测试：导出功能 (routers/export.py)"""
import pytest
from io import BytesIO
from openpyxl import load_workbook
from database import SessionLocal
from helpers import get_account_id
from factories import api_create_product, api_create_customer, api_create_supplier

from helpers import make_headers

HEADERS = make_headers()


def _count_excel_rows(content):
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    return ws.max_row


def _is_csv(content):
    text = content.decode("utf-8-sig")
    return "," in text


class TestExportProducts:
    def test_export_products_excel(self, client):
        api_create_product(client, HEADERS)
        resp = client.get("/api/export/products?format=excel", headers=HEADERS)
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert _count_excel_rows(resp.content) >= 2

    def test_export_products_csv(self, client):
        api_create_product(client, HEADERS)
        resp = client.get("/api/export/products?format=csv", headers=HEADERS)
        assert resp.status_code == 200
        assert "text/csv" in resp.headers["content-type"]
        assert _is_csv(resp.content)


class TestExportInventory:
    def test_export_inventory_excel(self, client):
        resp = client.get("/api/export/inventory?format=excel", headers=HEADERS)
        assert resp.status_code == 200

    def test_export_inventory_csv(self, client):
        resp = client.get("/api/export/inventory?format=csv", headers=HEADERS)
        assert resp.status_code == 200
        assert _is_csv(resp.content)

    def test_export_inventory_alert_only(self, client):
        resp = client.get("/api/export/inventory?format=excel&alert_only=true", headers=HEADERS)
        assert resp.status_code == 200


class TestExportPurchases:
    def test_export_purchases_excel(self, client):
        api_create_product(client, HEADERS)
        api_create_supplier(client, HEADERS)
        resp = client.get("/api/export/purchases?format=excel", headers=HEADERS)
        assert resp.status_code == 200

    def test_export_purchases_csv(self, client):
        resp = client.get("/api/export/purchases?format=csv", headers=HEADERS)
        assert resp.status_code == 200

    def test_export_purchases_with_date_filter(self, client):
        resp = client.get("/api/export/purchases?format=excel&start_date=2026-01-01&end_date=2026-12-31", headers=HEADERS)
        assert resp.status_code == 200


class TestExportSales:
    def test_export_sales_excel(self, client):
        api_create_product(client, HEADERS)
        api_create_customer(client, HEADERS)
        resp = client.get("/api/export/sales?format=excel", headers=HEADERS)
        assert resp.status_code == 200

    def test_export_sales_csv(self, client):
        resp = client.get("/api/export/sales?format=csv", headers=HEADERS)
        assert resp.status_code == 200

    def test_export_sales_with_date_filter(self, client):
        resp = client.get("/api/export/sales?format=excel&start_date=2026-01-01&end_date=2026-12-31", headers=HEADERS)
        assert resp.status_code == 200


class TestExportProductsBatch:
    def test_export_batch_valid_ids(self, client):
        pid, _ = api_create_product(client, HEADERS)
        resp = client.get(f"/api/export/products-batch?product_ids={pid}", headers=HEADERS)
        assert resp.status_code == 200

    def test_export_batch_empty_ids(self, client):
        resp = client.get("/api/export/products-batch?product_ids=", headers=HEADERS)
        assert resp.status_code in (400, 422)

    def test_export_batch_invalid_ids(self, client):
        resp = client.get("/api/export/products-batch?product_ids=abc", headers=HEADERS)
        assert resp.status_code in (400, 422)

    def test_export_batch_csv(self, client):
        pid, _ = api_create_product(client, HEADERS)
        resp = client.get(f"/api/export/products-batch?product_ids={pid}&format=csv", headers=HEADERS)
        assert resp.status_code == 200
        assert _is_csv(resp.content)


class TestExportProfit:
    def test_export_profit_excel(self, client):
        resp = client.get("/api/export/profit?format=excel", headers=HEADERS)
        assert resp.status_code == 200

    def test_export_profit_csv(self, client):
        resp = client.get("/api/export/profit?format=csv", headers=HEADERS)
        assert resp.status_code == 200
        assert _is_csv(resp.content)

    def test_export_profit_with_dates(self, client):
        resp = client.get("/api/export/profit?format=excel&start_date=2026-01-01&end_date=2026-12-31", headers=HEADERS)
        assert resp.status_code == 200
